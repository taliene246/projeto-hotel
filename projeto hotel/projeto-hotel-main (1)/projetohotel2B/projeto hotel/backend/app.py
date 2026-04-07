import os
from flask import Flask, request, jsonify, send_from_directory
import openpyxl  #Biblioteca para ler e escrever planilhas Excel (.xlsx)
from datetime import (
    datetime,
)  #Para registrar a data de cada cadastro automaticamente


# Caminho base do projeto (uma pasta acima do backend)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido
# Pasta frontend (HTML e JS)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\frontend
#
# Pasta static (CSS)
STATIC_DIR = os.path.join(BASE_DIR, "static")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\static

# Aqui definimos os diretórios e o nome do arquivo Excel que servirá de banco.
# =============================================================================
DB_DIR = os.path.join(
    os.path.dirname(__file__), "..", "db"
)  # Pasta onde ficará o banco de dados
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\db
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\db\clientes.xlsx

# Cabeçalhos das colunas do Excel (linha 1)
COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro",
]


# =============================================================================
# 🛠 FUNÇÃO AUXILIAR: CRIA O ARQUIVO EXCEL CASO NÃO EXISTA
# =============================================================================
# Essa função garante que o sistema funcione mesmo na primeira execução.
#    Se a pasta 'db' ou o arquivo 'clientes.xlsx' não existirem, eles são criados.
# =============================================================================
def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)  #Cria a pasta db se ainda não existir

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()  #Cria uma nova planilha Excel
        sheet = workbook.active  # Pega a planilha ativa
        sheet.title = "Clientes"  # Nomeia a aba principal
        sheet.append(COLUMNS)  # Adiciona os títulos das colunas
        workbook.save(EXCEL_FILE)  # Salva o arquivo Excel


app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/" + STATIC_DIR)


# =========================
# ROTA PRINCIPAL (HTML)
# ROTAS DAS PÁGINAS (FRONTEND)
# Página inicial (Cadastro/index.html)
# =========================
@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR, "index.html")


# Página de Consulta
@app.route("/consulta")
def consulta_page():
    return send_from_directory(FRONTEND_DIR, "consulta.html")


# Página de Alteração
@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")


#Rota para servir imagens, scripts ou outros arquivos na pasta “assets”
@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)


# -------------------------------------------------------------------------
#  CADASTRAR CLIENTE
# -------------------------------------------------------------------------
@app.route("/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """Recebe os dados do formulário (em JSON), valida e salva um novo cliente no Excel."""
    try:
        data = request.json  #Dados enviados do frontend via POST (JSON)
        #Campos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos.",
                    }
                ),
                400,
            )
        workbook = openpyxl.load_workbook(EXCEL_FILE)  #Abre o arquivo Excel
        sheet = workbook.active
        #Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        #Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""),  # Campo opcional
            datetime.now().strftime("%Y-%m-%d"),  # Data atual
        ]
        sheet.append(novo_cliente)  # Adiciona nova linha no Excel
        workbook.save(EXCEL_FILE)  #Salva alterações
        # Retorna mensagem de sucesso
        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cliente cadastrado com sucesso!",
                    "id": new_id,
                }
            ),
            201,
        )
    except Exception as e:
        # ⚠️ Tratamento de erro genérico
        return (
            jsonify({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,
        )
    
    @app.route("/api/buscar", methods=["GET"])
def buscar_clientes():
    """
    Busca clientes pelo nome (não diferencia maiúsculas/minúsculas).
    """
    nome_query = request.args.get("nome", "").lower()  # 🔤 Nome pesquisado

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        resultados = []  # 🧺 Lista para armazenar resultados

        # 🧭 Percorre todas as linhas (ignorando o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUMNS, row))  # Converte linha → dicionário
            nome_cliente = (cliente.get("Nome") or "").lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)  # 🔙 Retorna lista de clientes encontrados

    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado."}),
            404,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )


if __name__ == "__main__":
    print("BASE_DIR:", BASE_DIR)
    print("FRONTEND_DIR:", FRONTEND_DIR)
    print("STATIC_DIR:", STATIC_DIR)
    init_excel()
    app.run(debug=True)



