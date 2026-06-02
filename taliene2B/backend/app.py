import os
from flask import Flask, request, jsonify, send_from_directory
import openpyxl  # 📊 Biblioteca para ler e escrever planilhas Excel (.xlsx)
from datetime import (
    datetime,
)  # ⏰ Para registrar a data de cada cadastro automaticamente


# Caminho base do projeto (uma pasta acima do backend)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido
# Pasta frontend (HTML e JS)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\frontend
#
# Pasta static (CSS)
STATIC_DIR = os.path.join(BASE_DIR, "static")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\static

# 💾 Aqui definimos os diretórios e o nome do arquivo Excel que servirá de banco.
# =============================================================================
DB_DIR = os.path.join(
    os.path.dirname(__file__), "..", "db"
)  # Pasta onde ficará o banco de dados
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\db
EXCEL_FILE = os.path.join(DB_DIR, "clientes.xlsx")
# C:\Users\psicoton\Documents\PSS SEED\Logica de Programação\projeto_hotel_Guido\db\clientes.xlsx

# Cabeçalhos das colunas do Excel (linha 1)
COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro",
]


# =============================================================================
# 🛠 FUNÇÃO AUXILIAR: CRIA O ARQUIVO EXCEL CASO NÃO EXISTA
# =============================================================================
# 💡 Essa função garante que o sistema funcione mesmo na primeira execução.
#    Se a pasta 'db' ou o arquivo 'clientes.xlsx' não existirem, eles são criados.
# =============================================================================
def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)  # 🧱 Cria a pasta db se ainda não existir

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()  # 📗 Cria uma nova planilha Excel
        sheet = workbook.active  # Pega a planilha ativa
        sheet.title = "Clientes"  # Nomeia a aba principal
        sheet.append(COLUMNS)  # Adiciona os títulos das colunas
        workbook.save(EXCEL_FILE)  # Salva o arquivo Excel


app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="/static")


# =========================
# ROTA PRINCIPAL (HTML)
# ROTAS DAS PÁGINAS (FRONTEND)
# Página inicial (Cadastro/index.html)
# =========================
@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR, "index.html")


# Página de Consulta
@app.route("/consulta")
def consulta_page():
    return send_from_directory(FRONTEND_DIR, "consulta.html")


# Página de Alteração
@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")


# 📸 Rota para servir imagens, scripts ou outros arquivos na pasta “assets”
@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)


# -------------------------------------------------------------------------
# 📦 CADASTRAR CLIENTE
# -------------------------------------------------------------------------
@app.route("/api/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """Recebe os dados do formulário (em JSON), valida e salva um novo cliente no Excel."""
    try:
        data = request.json  # 📨 Dados enviados do frontend via POST (JSON)
        # ⚙️ Campos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos.",
                    }
                ),
                400,
            )
        workbook = openpyxl.load_workbook(EXCEL_FILE)  # 📂 Abre o arquivo Excel
        sheet = workbook.active
        # 🧮 Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        # 🧾 Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacao", ""),  # Campo opcional
            datetime.now().strftime("%Y-%m-%d"),  # 📅 Data atual
        ]
        sheet.append(novo_cliente)  # Adiciona nova linha no Excel
        workbook.save(EXCEL_FILE)  # 💾 Salva alterações
        # ✅ Retorna mensagem de sucesso
        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cliente cadastrado com sucesso!",
                    "id": new_id,
                }
            ),
            201,
        )
    except Exception as e:
        # ⚠️ Tratamento de erro genérico
        return (
            jsonify({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# 🔍 CONSULTAR CLIENTES por nome
# -------------------------------------------------------------------------
@app.route("/api/buscar", methods=["GET"])
def buscar_clientes():
    """
    Busca clientes pelo nome (não diferencia maiúsculas/minúsculas).
    """
    nome_query = request.args.get("nome", "").lower()  # 🔤 Nome pesquisado

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        resultados = []  # 🧺 Lista para armazenar resultados

        # 🧭 Percorre todas as linhas (ignorando o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUMNS, row))  # Converte linha → dicionário
            nome_cliente = (cliente.get("Nome") or "").lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)  # 🔙 Retorna lista de clientes encontrados

    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado."}),
            404,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# 📋 CONSULTAR CLIENTE POR ID
# -------------------------------------------------------------------------
@app.route("/api/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):
    """
    Retorna os dados completos de um cliente pelo seu ID.
    """
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # 🔍 Procura o cliente linha por linha
        for row_idx in range(2, sheet.max_row + 1):
            row_id = sheet.cell(row=row_idx, column=1).value
            if row_id == cliente_id:
                row_values = [cell.value for cell in sheet[row_idx]]
                cliente = dict(zip(COLUMNS, row_values))
                return jsonify(cliente)

        # ❌ Se não encontrar o ID
        return jsonify({"status": "error", "message": "Cliente não encontrado."}), 404
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao buscar cliente: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# ✏️ ATUALIZAR DADOS DE UM CLIENTE
# -------------------------------------------------------------------------
# 1. Definimos o "endereço" (rota) onde o navegador ou app deve bater para atualizar um cliente.
# O <int:cliente_id> é o número único do hóspede que queremos encontrar.
# POST indica que estamos enviando novos dados para o servidor.
@app.route("/api/atualizar/<int:cliente_id>", methods=["POST"])
def atualizar_cliente(cliente_id):
    """
    Função para alterar as informações de um hóspede no nosso banco de dados (Excel).
    """
    try:
        # Pega o "pacote de dados" (JSON - JavaScript Object Notation) que veio da internet.
        # Nele estão o novo nome, telefone, etc., que o usuário digitou no site.
        data = request.json

        # Abre o nosso arquivo Excel (que funciona como o cérebro do hotel para guardar dados).
        workbook = openpyxl.load_workbook(EXCEL_FILE)

        # Seleciona a página que está aberta/ativa no momento dentro do arquivo.
        sheet = workbook.active
        # 🧭 BUSCA: Precisamos achar em qual linha do Excel este hóspede está.
        # Começamos com -1 para indicar que, por enquanto, não o encontramos.
        row_to_update = -1
        # O sistema começa a ler da linha 2 (pulando os títulos) até a última linha escrita.
        for row_idx in range(2, sheet.max_row + 1):
            # Se o valor da primeira coluna (ID) for igual ao ID que recebemos, achamos!
            if sheet.cell(row=row_idx, column=1).value == cliente_id:
                row_to_update = row_idx  # Guardamos o número da linha.
                break  # Para de procurar para economizar processamento.
        # Se depois de ler tudo continuarmos com -1, o hóspede não existe.
        if row_to_update == -1:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Cliente não encontrado para atualização.",
                    }
                ),
                404,  # Código de erro padrão para "Não Encontrado".
            )
        # 🧾 ESCRITA: Agora o sistema escreve as novas informações nas colunas certas daquela linha.
        # Column 2 = Nome, Column 3 = CPF, e assim por diante.
        sheet.cell(row=row_to_update, column=2, value=data.get("nome"))
        sheet.cell(row=row_to_update, column=3, value=data.get("cpf"))
        sheet.cell(row=row_to_update, column=4, value=data.get("email"))
        sheet.cell(row=row_to_update, column=5, value=data.get("telefone"))
        sheet.cell(row=row_to_update, column=6, value=data.get("endereco"))
        sheet.cell(row=row_to_update, column=7, value=data.get("observacao"))

        # 💾 SALVAR: Importante! Se não salvar, as mudanças ficam só na memória e somem depois.
        workbook.save(EXCEL_FILE)

        # Retorna uma mensagem de sucesso para o Front-end mostrar na tela do usuário.
        return jsonify(
            {
                "status": "success",
                "message": "Dados do cliente atualizados com sucesso!",
            }
        )

    except Exception as e:
        # Se houver algum erro inesperado (ex: o arquivo Excel estar aberto por outra pessoa),
        # o sistema "captura" o erro e avisa o que aconteceu, sem travar o programa todo.
        return (
            jsonify({"status": "error", "message": f"Erro ao atualizar dados: {e}"}),
            500,  # Código de erro para "Falha Interna no Servidor".
        )


if __name__ == "__main__":
    print("BASE_DIR:", BASE_DIR)
    print("FRONTEND_DIR:", FRONTEND_DIR)
    print("STATIC_DIR:", STATIC_DIR)
    init_excel()
    app.run(
        debug=True, port=5000
    )  # 🚀 Inicia o servidor local em http://localhost:5000
