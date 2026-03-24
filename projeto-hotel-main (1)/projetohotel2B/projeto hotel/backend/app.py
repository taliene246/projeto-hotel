import os
from flask import (
    Flask,
    request,
    jsonify,
    send_from_directory,
)
import os
import openpyxl  # 📊 Biblioteca para ler e escrever planilhas Excel (.xlsx)
from datetime import (
    datetime,
)

#caminho base do projeto (uma pasta acima de backend)
BASE_DIR = os.path.abspath (os.path.join(os.path.dirname(__file__), ".."))

#pasta frondend (HTML e JS)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

#pasta static (css)
STATIC_DIR = os.path.join(BASE_DIR,"static")

DB_DIR = os.path.join(os.path.dirname(__file__),"..","db")
EXCEL_FILE = os.path.join(DB_DIR,"clientes.xlsx")

#cabeçalhos das colunas do excel (linha1)
COLUMNS = [
   "id",
   "nome",
   "cpf",
   "email",
   "telefone",
   "endereço",
   "observaçãoes",
   "data cadastro",
]




app = Flask(__name__,static_folder=STATIC_DIR,static_url_path="/"+STATIC_DIR)

@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR,"index.html")


@app.route("/consulta")
def consulta():
    return send_from_directory(FRONTEND_DIR,"consulta.html")

@app.route("/alterar")
def alterar_pege():
    return send_from_directory(FRONTEND_DIR,"alterar.html")

@app.route("/assets/<path:filename>")
def assets(filename): 
   return send_from_directory("../frontend/assets", filename)

@app.route("/api/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """
    Recebe os dados do formulário (em JSON), valida e salva um novo cliente no Excel.
    """
    try:
        data = request.json  # 📨 Dados enviados do frontend via POST (JSON)

        # ⚙️ Campos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos.",
                    }
                ),
                400,
            )

        workbook = openpyxl.load_workbook(EXCEL_FILE)  # 📂 Abre o arquivo Excel
        sheet = workbook.active

        # 🧮 Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        # 🧾 Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""),  # Campo opcional
            datetime.now().strftime("%Y-%m-%d"),  # 📅 Data atual
        ]
        worbook = openpyxl.load_worbook(EXCEL_FILE)  
      sheet = worbook.active 
      last_id = 0    
      if sheet.max_row > 1:   
            last_id = sheet.cell(row=sheet.max_row, column=1). value or 0 
            new_id = last_id + 1    
      novo_cliente =[      
            new_id,   
            data.get("nome"),    
            data.get("cpf"),    
            data.get("email"),   
            data.get("telefone"),
      ]
      sheet.append(novo_cliente)
      worbook.save(EXCEL_FILE)
      return (
         jsonify(|
            {
                  "status":"succes",
                  "massage": "cliente cadastro com sucesso!",
                  "id": new_id,
            }
         )
         201
      )
   except Exception as e:
   return(
      jsonify({"status": "error", "mensage": f"erro ao salvar no servidor: {e}"})
      500,
    ),
      




if __name__ == "__main__":
 print("BASE_DIR:", BASE_DIR)
 print("FRONTEND_DIR:", STATIC_DIR)
 print("STATIC_DIR:",STATIC_DIR)
app.run(debug=True)


